"""
CUBOX 2.0 - Fase 3B: import Excel profile-aware (BOX/PALLET/PANEL/CUSTOM) +
generador de templates. Cubre los Tests A-Q especificados para esta fase.

El importador legacy (POST /api/import-excel, core/excel_import.py) no se
toca -Test A solo confirma que sigue funcionando igual."""

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.main import app

client = TestClient(app)
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_workbook_bytes(headers: list[str], rows: list[list]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _import(profile: str, headers: list[str], rows: list[list]):
    content = _build_workbook_bytes(headers, rows)
    r = client.post(
        "/api/import-items-excel",
        files={"file": (f"{profile}.xlsx", content, XLSX_MIME)},
        data={"profile": profile},
    )
    return r


# ---------------------------------------------------------------------------
# TEST A - el importador legacy sigue funcionando exactamente igual.
# ---------------------------------------------------------------------------


def test_legacy_import_excel_still_works():
    content = _build_workbook_bytes(
        ["Code", "Width", "Height", "Thickness", "Weight", "Quantity"],
        [["W1", 1200, 2000, 100, 45, 3]],
    )
    r = client.post("/api/import-excel", files={"file": ("legacy.xlsx", content, XLSX_MIME)})
    assert r.status_code == 200
    items = r.json()
    assert len(items) == 1
    assert items[0]["code"] == "W1"
    assert items[0]["item_type"] == "panel"
    assert items[0]["dimensions"] == {"length": 1200, "width": 100, "height": 2000}


# ---------------------------------------------------------------------------
# TEST B - import de BOX: ItemType.BOX, Dimensions3D exacto, FREE por defecto.
# ---------------------------------------------------------------------------


def test_box_import_default_free_orientation():
    r = _import("box", ["Code", "Quantity", "Length", "Width", "Height", "Weight"], [["BOX1", 5, 600, 400, 300, 25]])
    assert r.status_code == 200
    preview = r.json()
    assert preview["is_valid"] is True
    assert len(preview["items"]) == 1
    item = preview["items"][0]
    assert item["item_type"] == "box"
    assert item["dimensions"] == {"length": 600, "width": 400, "height": 300}
    assert item["orientation_policy"] == "free"


# ---------------------------------------------------------------------------
# TEST C - BOX con Orientation=UPRIGHT.
# ---------------------------------------------------------------------------


def test_box_import_upright_orientation():
    r = _import(
        "box",
        ["Code", "Quantity", "Length", "Width", "Height", "Weight", "Orientation"],
        [["BOX2", 2, 600, 400, 300, 25, "upright"]],
    )
    preview = r.json()
    assert preview["is_valid"] is True
    assert preview["items"][0]["orientation_policy"] == "upright"


# ---------------------------------------------------------------------------
# TEST D - import de PALLET: ItemType.PALLET, UPRIGHT, dimensiones intactas.
# ---------------------------------------------------------------------------


def test_pallet_import_defaults_to_upright():
    r = _import("pallet", ["Code", "Quantity", "Length", "Width", "Height", "Weight"], [["PAL1", 4, 1200, 1000, 150, 20]])
    preview = r.json()
    assert preview["is_valid"] is True
    item = preview["items"][0]
    assert item["item_type"] == "pallet"
    assert item["orientation_policy"] == "upright"
    assert item["dimensions"] == {"length": 1200, "width": 1000, "height": 150}


# ---------------------------------------------------------------------------
# TEST E - import de PANEL: mapeo canonico correcto, regla de vidrio.
# ---------------------------------------------------------------------------


def test_panel_import_preserves_glass_face_rule():
    r = _import("panel", ["Code", "Quantity", "Width", "Height", "Thickness", "Weight"], [["W1", 3, 1200, 2000, 100, 45]])
    preview = r.json()
    assert preview["is_valid"] is True
    item = preview["items"][0]
    assert item["item_type"] == "panel"
    assert item["orientation_policy"] == "panel_edge_only"
    assert item["dimensions"] == {"length": 1200, "width": 100, "height": 2000}


# ---------------------------------------------------------------------------
# TEST F - import de CUSTOM: Orientation explicito obligatorio.
# ---------------------------------------------------------------------------


def test_custom_import_uses_explicit_orientation():
    r = _import(
        "custom",
        ["Code", "Quantity", "Length", "Width", "Height", "Weight", "Orientation"],
        [["C1", 1, 500, 500, 500, 10, "fixed"]],
    )
    preview = r.json()
    assert preview["is_valid"] is True
    item = preview["items"][0]
    assert item["item_type"] == "custom"
    assert item["orientation_policy"] == "fixed"


def test_custom_import_missing_orientation_column_is_workbook_level_error():
    """Si la columna Orientation no existe en absoluto, se reporta a nivel
    de archivo (igual que cualquier otra columna obligatoria ausente -ver
    Test G)."""
    r = _import("custom", ["Code", "Quantity", "Length", "Width", "Height", "Weight"], [["C2", 1, 500, 500, 500, 10]])
    preview = r.json()
    assert preview["is_valid"] is False
    error = preview["errors"][0]
    assert error["row"] is None
    assert error["code"] == "MISSING_REQUIRED_COLUMN"
    assert "orientation" in error["message"].lower()


def test_custom_import_empty_orientation_value_is_row_level_error():
    """Si la columna existe pero una fila deja el valor vacio, se reporta
    a nivel de esa fila."""
    r = _import(
        "custom",
        ["Code", "Quantity", "Length", "Width", "Height", "Weight", "Orientation"],
        [["C2", 1, 500, 500, 500, 10, ""]],
    )
    preview = r.json()
    assert preview["is_valid"] is False
    assert any(e["code"] == "MISSING_REQUIRED_VALUE" and e["column"] == "orientation" for e in preview["errors"])


def test_custom_import_rejects_panel_edge_only():
    r = _import(
        "custom",
        ["Code", "Quantity", "Length", "Width", "Height", "Weight", "Orientation"],
        [["C3", 1, 500, 500, 500, 10, "PANEL_EDGE_ONLY"]],
    )
    preview = r.json()
    assert preview["is_valid"] is False
    assert any(e["code"] == "UNSUPPORTED_ORIENTATION_FOR_PROFILE" for e in preview["errors"])


# ---------------------------------------------------------------------------
# TEST G - columna obligatoria faltante -> error estructurado a nivel de
# archivo.
# ---------------------------------------------------------------------------


def test_missing_required_column_reports_structured_error():
    r = _import("box", ["Code", "Quantity", "Length", "Width", "Weight"], [["BOX1", 1, 600, 400, 25]])  # falta Height
    preview = r.json()
    assert preview["is_valid"] is False
    assert len(preview["errors"]) == 1
    error = preview["errors"][0]
    assert error["row"] is None
    assert error["code"] == "MISSING_REQUIRED_COLUMN"
    assert "height" in error["message"].lower()


# ---------------------------------------------------------------------------
# TEST H - multiples filas invalidas: todos los errores en un solo preview.
# ---------------------------------------------------------------------------


def test_multiple_invalid_rows_all_reported_together():
    r = _import(
        "box",
        ["Code", "Quantity", "Length", "Width", "Height", "Weight"],
        [
            ["BOX1", "not-a-number", 600, 400, 300, 25],
            ["BOX2", 2, 600, 400, "abc", 25],
            ["BOX3", 2, 600, 400, 300, -5],
            ["BOX4", 2, 600, 400, 300, 25],
        ],
    )
    preview = r.json()
    assert preview["is_valid"] is False
    assert len(preview["items"]) == 1
    assert preview["items"][0]["code"] == "BOX4"
    assert {e["row"] for e in preview["errors"]} == {2, 3, 4}


# ---------------------------------------------------------------------------
# TEST I - booleano invalido.
# ---------------------------------------------------------------------------


def test_invalid_stackable_value_reports_structured_error():
    r = _import(
        "box",
        ["Code", "Quantity", "Length", "Width", "Height", "Weight", "Stackable"],
        [["BOX1", 1, 600, 400, 300, 25, "maybe"]],
    )
    preview = r.json()
    assert preview["is_valid"] is False
    assert any(e["code"] == "INVALID_BOOLEAN" and e["column"] == "stackable" for e in preview["errors"])


# ---------------------------------------------------------------------------
# TEST J - orientacion invalida.
# ---------------------------------------------------------------------------


def test_invalid_orientation_value_reports_structured_error():
    r = _import(
        "box",
        ["Code", "Quantity", "Length", "Width", "Height", "Weight", "Orientation"],
        [["BOX1", 1, 600, 400, 300, 25, "SIDEWAYS"]],
    )
    preview = r.json()
    assert preview["is_valid"] is False
    assert any(e["code"] == "INVALID_ORIENTATION" for e in preview["errors"])


# ---------------------------------------------------------------------------
# TEST K - default seguro de Stackable (BOX/PALLET/CUSTOM = No + warning;
# PANEL conserva el default legacy = Si, sin warning).
# ---------------------------------------------------------------------------


def test_box_without_stackable_defaults_to_false_with_warning():
    r = _import("box", ["Code", "Quantity", "Length", "Width", "Height", "Weight"], [["BOX1", 1, 600, 400, 300, 25]])
    preview = r.json()
    assert preview["is_valid"] is True
    assert preview["items"][0]["stackable"] is False
    assert any(w["code"] == "STACKABLE_DEFAULTED" for w in preview["warnings"])


def test_panel_without_stackable_keeps_legacy_default_true():
    r = _import("panel", ["Code", "Quantity", "Width", "Height", "Thickness", "Weight"], [["W1", 1, 1200, 2000, 100, 45]])
    preview = r.json()
    assert preview["is_valid"] is True
    assert preview["items"][0]["stackable"] is True
    assert preview["warnings"] == []


# ---------------------------------------------------------------------------
# TEST L - resumen: peso total = Sum(quantity * unit weight).
# ---------------------------------------------------------------------------


def test_summary_calculates_total_weight_as_quantity_times_unit_weight():
    r = _import(
        "box",
        ["Code", "Quantity", "Length", "Width", "Height", "Weight"],
        [["BOX1", 5, 600, 400, 300, 25], ["BOX2", 2, 500, 500, 500, 10]],
    )
    preview = r.json()
    assert preview["summary"]["total_units"] == 7
    assert preview["summary"]["total_weight"] == pytest.approx(5 * 25 + 2 * 10)
    assert preview["summary"]["valid_rows"] == 2
    assert preview["summary"]["total_rows"] == 2
    assert preview["summary"]["unique_codes"] == 2


# ---------------------------------------------------------------------------
# TEST M/N/O/P - templates por perfil.
# ---------------------------------------------------------------------------


def _template_headers(profile: str) -> list[str]:
    r = client.get(f"/api/import-template/{profile}")
    assert r.status_code == 200
    workbook = load_workbook(io.BytesIO(r.content))
    assert set(workbook.sheetnames) == {"Items", "Instructions"}
    return [c.value for c in next(workbook["Items"].iter_rows(min_row=1, max_row=1))]


def test_box_template_has_expected_headers():
    headers = _template_headers("box")
    assert headers == [
        "Code", "Quantity", "Length", "Width", "Height", "Weight",
        "Description", "Orientation", "Stackable", "Max Stack Weight", "Group", "Priority", "Delivery Sequence",
    ]


def test_pallet_template_has_optional_orientation_column():
    """Fase 6: PALLET ahora SI expone Orientation (opcional, UPRIGHT/FIXED
    unicamente) -antes no tenia esta columna, ver import_items.py seccion 8
    del pedido de Fase 6 para la justificacion de este cambio."""
    headers = _template_headers("pallet")
    assert "Orientation" in headers
    assert headers[:6] == ["Code", "Quantity", "Length", "Width", "Height", "Weight"]


def test_panel_template_uses_width_height_thickness():
    headers = _template_headers("panel")
    assert "Thickness" in headers
    assert "Length" not in headers
    assert "System" in headers


def test_custom_template_requires_orientation_column():
    headers = _template_headers("custom")
    assert "Orientation" in headers


# ---------------------------------------------------------------------------
# TEST Q - round trip: generar template, poblarlo, importarlo.
# ---------------------------------------------------------------------------


def test_round_trip_template_to_import():
    r = client.get("/api/import-template/box")
    workbook = load_workbook(io.BytesIO(r.content))
    items_sheet = workbook["Items"]
    items_sheet.append(["RT1", 3, 700, 500, 400, 30, "Round trip test", "UPRIGHT", "Yes", "", "G1", 2, ""])
    buffer = io.BytesIO()
    workbook.save(buffer)

    r2 = client.post(
        "/api/import-items-excel", files={"file": ("box.xlsx", buffer.getvalue(), XLSX_MIME)}, data={"profile": "box"}
    )
    preview = r2.json()
    assert preview["is_valid"] is True
    item = preview["items"][0]
    assert item["code"] == "RT1"
    assert item["dimensions"] == {"length": 700, "width": 500, "height": 400}
    assert item["orientation_policy"] == "upright"
    assert item["stackable"] is True
    assert item["group"] == "G1"
    assert item["priority"] == 2


# ---------------------------------------------------------------------------
# Sanidad adicional: el preview nunca muta el estado activo de cubicaje.
# ---------------------------------------------------------------------------


def test_import_preview_does_not_mutate_active_state():
    """El estado activo (`_current_state`, compartido por todo el proceso)
    puede tener o no un cubicaje segun que otros tests hayan corrido antes
    -lo que importa es que /api/import-items-excel no lo cambie en absoluto,
    sin importar cual sea su valor de partida."""
    before = client.get("/api/state")

    r = _import("box", ["Code", "Quantity", "Length", "Width", "Height", "Weight"], [["BOX1", 1, 600, 400, 300, 25]])
    assert r.json()["is_valid"] is True

    after = client.get("/api/state")
    assert after.status_code == before.status_code
    assert after.json() == before.json()
