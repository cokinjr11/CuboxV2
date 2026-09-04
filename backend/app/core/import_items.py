"""Import Excel CUBOX 2.0, profile-aware (Fase 3B).

No reemplaza a excel_import.py (legacy, sigue exactamente igual para
POST /api/import-excel) -este modulo alimenta el nuevo
POST /api/import-items-excel, que recibe un `profile` (BOX/PALLET/PANEL/
CUSTOM, reutilizando ItemType -no hace falta un enum ImportProfile
separado) y devuelve un ImportPreview: parsea la hoja completa, junta TODOS
los errores/warnings encontrados (no se detiene en la primera fila
invalida) y nunca muta el estado activo de cubicaje.

Mapeo de dimensiones -UNA sola implementacion, reutilizada, nunca duplicada:
  - PANEL: Width/Height/Thickness -> dimensions_from_legacy() (models/schemas.py),
    la misma funcion que usa el importador legacy y toda la Fase 3A/3A.1.
  - BOX/PALLET/CUSTOM: Length/Width/Height -> Dimensions3D(length, width, height)
    directo, SIN pasar por la semantica de panel.
"""

import io
from dataclasses import dataclass

from openpyxl import load_workbook

from app.core.excel_import import TRUE_VALUES
from app.models.import_schemas import ImportDefaults, ImportIssue, ImportIssueSeverity, ImportPreview, ImportSummary
from app.models.schemas import Dimensions3D, ItemType, LoadItem, OrientationPolicy, dimensions_from_legacy

FALSE_VALUES = {"no", "n", "false", "0"}
"""Complemento explicito de TRUE_VALUES (importado de excel_import.py, no
duplicado): a diferencia del importador legacy -que trata cualquier valor
no reconocido como False silenciosamente-, el import profile-aware
distingue un False real de un valor invalido (ver TEST I de esta fase)."""


_COLUMN_ALIASES: dict[str, str] = {
    "code": "code",
    "description": "description",
    "length": "length",
    "width": "width",
    "height": "height",
    "thickness": "thickness",
    "weight": "weight",
    "quantity": "quantity",
    "qty": "quantity",
    "system": "system",
    "group": "group",
    "stackable": "stackable",
    "priority": "priority",
    "maxstackweight": "max_stack_weight",
    "deliverysequence": "delivery_sequence",
    "stop": "delivery_sequence",
    "orientation": "orientation",
}

_ORIENTATION_ALIASES: dict[str, OrientationPolicy] = {
    "free": OrientationPolicy.FREE,
    "upright": OrientationPolicy.UPRIGHT,
    "fixed": OrientationPolicy.FIXED,
    "paneledgeonly": OrientationPolicy.PANEL_EDGE_ONLY,
}


def _normalize_header(value) -> str:
    """Case-insensitive y ademas ignora espacios/guiones bajos, para que
    "Max Stack Weight", "MaxStackWeight" y "max_stack_weight" normalicen
    todos a la misma clave -el importador legacy solo hace lower(), esto es
    deliberadamente mas tolerante (el template nuevo usa headers con
    espacios; ver core/import_templates.py)."""
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "").replace("_", "")


def _parse_boolean_or_none(value) -> bool | None:
    """None si el valor no es reconocible como booleano -a diferencia de
    excel_import._parse_stackable, que nunca falla (trata todo lo
    desconocido como False). Centraliza el parseo: una sola funcion para
    los 4 perfiles (BOX/PALLET/PANEL/CUSTOM), ver seccion 18 del plan."""
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    return None


def _parse_orientation(value) -> OrientationPolicy | None:
    return _ORIENTATION_ALIASES.get(_normalize_header(value))


@dataclass(frozen=True)
class _ProfileSpec:
    required: tuple[str, ...]
    dimension_keys: tuple[str, str, str]
    use_legacy_panel_mapping: bool
    item_type: ItemType
    orientation_mode: str  # "none" | "optional" | "required"
    allowed_orientations: tuple[OrientationPolicy, ...]
    default_orientation: OrientationPolicy | None
    legacy_stackable_default: bool  # True = default Si (perfil PANEL, igual que el importador legacy)
    allow_system: bool


PROFILE_SPECS: dict[ItemType, _ProfileSpec] = {
    ItemType.BOX: _ProfileSpec(
        required=("code", "quantity", "length", "width", "height", "weight"),
        dimension_keys=("length", "width", "height"),
        use_legacy_panel_mapping=False,
        item_type=ItemType.BOX,
        orientation_mode="optional",
        allowed_orientations=(OrientationPolicy.FREE, OrientationPolicy.UPRIGHT, OrientationPolicy.FIXED),
        default_orientation=OrientationPolicy.FREE,
        legacy_stackable_default=False,
        allow_system=False,
    ),
    ItemType.PALLET: _ProfileSpec(
        # Fase 6: Orientation es OPCIONAL para PALLET (no "none" como antes) -
        # reusa exactamente el mismo mecanismo generico de precedencia que ya
        # usa BOX (Excel explicito > default del plan/Floor Rotation > UPRIGHT
        # de sistema), solo que restringido a UPRIGHT/FIXED (un pallet nunca
        # es FREE ni PANEL_EDGE_ONLY). Si la celda viene vacia y no hay
        # default de plan, cae en default_orientation=UPRIGHT (igual que
        # siempre).
        required=("code", "quantity", "length", "width", "height", "weight"),
        dimension_keys=("length", "width", "height"),
        use_legacy_panel_mapping=False,
        item_type=ItemType.PALLET,
        orientation_mode="optional",
        allowed_orientations=(OrientationPolicy.UPRIGHT, OrientationPolicy.FIXED),
        default_orientation=OrientationPolicy.UPRIGHT,
        legacy_stackable_default=False,
        allow_system=False,
    ),
    ItemType.PANEL: _ProfileSpec(
        required=("code", "quantity", "width", "height", "thickness", "weight"),
        dimension_keys=("width", "height", "thickness"),
        use_legacy_panel_mapping=True,
        item_type=ItemType.PANEL,
        orientation_mode="none",
        allowed_orientations=(),
        default_orientation=OrientationPolicy.PANEL_EDGE_ONLY,
        legacy_stackable_default=True,
        allow_system=True,
    ),
    ItemType.CUSTOM: _ProfileSpec(
        required=("code", "quantity", "length", "width", "height", "weight", "orientation"),
        dimension_keys=("length", "width", "height"),
        use_legacy_panel_mapping=False,
        item_type=ItemType.CUSTOM,
        orientation_mode="required",
        allowed_orientations=(OrientationPolicy.FREE, OrientationPolicy.UPRIGHT, OrientationPolicy.FIXED),
        default_orientation=None,
        legacy_stackable_default=False,
        allow_system=False,
    ),
}


def _empty_summary() -> ImportSummary:
    return ImportSummary(total_rows=0, valid_rows=0, invalid_rows=0, total_units=0, total_weight=0.0, unique_codes=0)


def _parse_row(
    profile: ItemType,
    spec: _ProfileSpec,
    headers: list[str],
    row: tuple,
    row_number: int,
    defaults: ImportDefaults | None = None,
) -> tuple[LoadItem | None, list[ImportIssue], list[ImportIssue]]:
    values = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
    errors: list[ImportIssue] = []
    warnings: list[ImportIssue] = []

    def err(column: str, code: str, message: str) -> None:
        errors.append(ImportIssue(row=row_number, column=column, code=code, message=message, severity=ImportIssueSeverity.ERROR))

    def warn(column: str, code: str, message: str) -> None:
        warnings.append(
            ImportIssue(row=row_number, column=column, code=code, message=message, severity=ImportIssueSeverity.WARNING)
        )

    # "orientation" se excluye aca a proposito: cuando orientation_mode ==
    # "required" (CUSTOM), el bloque de orientacion mas abajo ya emite su
    # propio MISSING_REQUIRED_VALUE -evita reportar el mismo problema 2 veces.
    for missing_col in [c for c in spec.required if c != "orientation" and values.get(c) in (None, "")]:
        err(missing_col, "MISSING_REQUIRED_VALUE", f"Falta un valor para la columna obligatoria '{missing_col}'")

    code = str(values.get("code") or "").strip()

    quantity = None
    if values.get("quantity") not in (None, ""):
        try:
            quantity = int(values["quantity"])
            if quantity <= 0:
                raise ValueError
        except (TypeError, ValueError):
            err("quantity", "INVALID_NUMBER", "Quantity debe ser un entero mayor a 0")
            quantity = None

    weight = None
    if values.get("weight") not in (None, ""):
        try:
            weight = float(values["weight"])
            if weight <= 0:
                raise ValueError
        except (TypeError, ValueError):
            err("weight", "INVALID_NUMBER", "Weight debe ser un numero mayor a 0")
            weight = None

    dim_values: dict[str, float] = {}
    for key in spec.dimension_keys:
        if values.get(key) in (None, ""):
            continue  # ya se reporto arriba si era obligatoria
        try:
            v = float(values[key])
            if v <= 0:
                raise ValueError
            dim_values[key] = v
        except (TypeError, ValueError):
            err(key, "INVALID_NUMBER", f"{key.capitalize()} debe ser un numero mayor a 0")

    dimensions: Dimensions3D | None = None
    if len(dim_values) == len(spec.dimension_keys):
        if spec.use_legacy_panel_mapping:
            dimensions = dimensions_from_legacy(dim_values["width"], dim_values["height"], dim_values["thickness"])
        else:
            dimensions = Dimensions3D(length=dim_values["length"], width=dim_values["width"], height=dim_values["height"])

    # Precedencia (Fase 5): valor explicito de Excel > default del PLAN
    # (ImportDefaults, elegido en Handling Rules) > default de SISTEMA
    # (spec.legacy_stackable_default). El warning describe el default que
    # REALMENTE se aplico -nunca uno generico que no refleje la fuente real.
    raw_stackable = values.get("stackable")
    if raw_stackable in (None, ""):
        if defaults is not None and defaults.stackable is not None:
            stackable = defaults.stackable
            warn("stackable", "STACKABLE_DEFAULTED", f"Stackable no fue especificado; default del plan aplicado ({'Yes' if stackable else 'No'})")
        else:
            stackable = spec.legacy_stackable_default
            if not spec.legacy_stackable_default:
                warn("stackable", "STACKABLE_DEFAULTED", "Stackable no fue especificado; se asumio No")
    else:
        parsed_bool = _parse_boolean_or_none(raw_stackable)
        if parsed_bool is None:
            err("stackable", "INVALID_BOOLEAN", f"Valor de Stackable no reconocido: '{raw_stackable}'")
            stackable = spec.legacy_stackable_default
        else:
            stackable = parsed_bool

    # Misma precedencia para Orientation, pero el default del plan solo
    # aplica cuando el perfil tiene una columna de Orientation real
    # (orientation_mode != "none" -PALLET/PANEL tienen su politica fija, un
    # default de plan no tendria sentido ahi).
    orientation_policy = spec.default_orientation
    if spec.orientation_mode != "none":
        raw_orientation = values.get("orientation")
        if raw_orientation in (None, ""):
            if defaults is not None and defaults.orientation_policy is not None:
                orientation_policy = defaults.orientation_policy
                warn(
                    "orientation",
                    "ORIENTATION_DEFAULTED",
                    f"Orientation no fue especificado; default del plan aplicado ({orientation_policy.value})",
                )
            elif spec.orientation_mode == "required":
                err("orientation", "MISSING_REQUIRED_VALUE", "Falta un valor para la columna obligatoria 'orientation'")
        else:
            parsed_policy = _parse_orientation(raw_orientation)
            if parsed_policy is None:
                err("orientation", "INVALID_ORIENTATION", f"Valor de Orientation no reconocido: '{raw_orientation}'")
            elif parsed_policy not in spec.allowed_orientations:
                err(
                    "orientation",
                    "UNSUPPORTED_ORIENTATION_FOR_PROFILE",
                    f"Orientation '{raw_orientation}' no esta soportado para el perfil {profile.value}",
                )
            else:
                orientation_policy = parsed_policy

    priority = 0
    if values.get("priority") not in (None, ""):
        try:
            priority = int(values["priority"])
        except (TypeError, ValueError):
            err("priority", "INVALID_NUMBER", "Priority debe ser un numero entero")

    max_stack_weight = None
    if values.get("max_stack_weight") not in (None, ""):
        try:
            max_stack_weight = float(values["max_stack_weight"])
        except (TypeError, ValueError):
            err("max_stack_weight", "INVALID_NUMBER", "Max Stack Weight debe ser numerico")

    delivery_sequence = None
    if values.get("delivery_sequence") not in (None, ""):
        try:
            delivery_sequence = int(values["delivery_sequence"])
        except (TypeError, ValueError):
            err("delivery_sequence", "INVALID_NUMBER", "Delivery Sequence debe ser un numero entero")

    if errors or dimensions is None or not code or quantity is None or weight is None:
        return None, errors, warnings

    item = LoadItem(
        code=code,
        description=str(values.get("description") or "").strip(),
        dimensions=dimensions,
        weight=weight,
        quantity=quantity,
        system=str(values.get("system") or "").strip() if spec.allow_system else "",
        group=str(values.get("group") or "").strip(),
        stackable=stackable,
        priority=priority,
        max_stack_weight=max_stack_weight,
        delivery_sequence=delivery_sequence,
        item_type=spec.item_type,
        orientation_policy=orientation_policy,
    )
    return item, errors, warnings


def build_import_preview(file_bytes: bytes, profile: ItemType, defaults: ImportDefaults | None = None) -> ImportPreview:
    """Parsea la hoja completa para `profile` y devuelve un ImportPreview.
    NO muta ningun estado -solo lectura/validacion (ver Fase 3B, seccion 30).

    `defaults` (Fase 5) son los defaults del PLAN elegidos en Handling
    Rules -se aplican solo cuando la celda de Excel viene vacia, nunca
    pisan un valor explicito (ver _parse_row). El ImportPreview resultante
    ya contiene los LoadItems FINALES: no hay ningun merge oculto despues
    de esto (Fase 5, seccion 8) -lo que el usuario ve en Review es
    exactamente lo que entra al workspace."""
    spec = PROFILE_SPECS[profile]

    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        issue = ImportIssue(
            code="INVALID_FILE_FORMAT", message="El archivo no es un .xlsx valido", severity=ImportIssueSeverity.ERROR
        )
        return ImportPreview(profile=profile, is_valid=False, items=[], errors=[issue], warnings=[], summary=_empty_summary())

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        issue = ImportIssue(code="EMPTY_SHEET", message="El archivo Excel esta vacio", severity=ImportIssueSeverity.ERROR)
        return ImportPreview(profile=profile, is_valid=False, items=[], errors=[issue], warnings=[], summary=_empty_summary())

    headers = [_COLUMN_ALIASES.get(_normalize_header(h), _normalize_header(h)) for h in header_row]

    missing_columns = [c for c in spec.required if c not in headers]
    if missing_columns:
        issue = ImportIssue(
            code="MISSING_REQUIRED_COLUMN",
            message=f"Faltan columnas obligatorias en el Excel: {', '.join(missing_columns)}",
            severity=ImportIssueSeverity.ERROR,
        )
        return ImportPreview(profile=profile, is_valid=False, items=[], errors=[issue], warnings=[], summary=_empty_summary())

    items: list[LoadItem] = []
    errors: list[ImportIssue] = []
    warnings: list[ImportIssue] = []
    seen_codes: set[str] = set()
    total_rows = 0

    for row_number, row in enumerate(rows_iter, start=2):
        if row is None or all(v is None for v in row):
            continue
        total_rows += 1
        item, row_errors, row_warnings = _parse_row(profile, spec, headers, row, row_number, defaults)
        errors.extend(row_errors)
        warnings.extend(row_warnings)
        if item is not None:
            items.append(item)
            seen_codes.add(item.code)

    summary = ImportSummary(
        total_rows=total_rows,
        valid_rows=len(items),
        invalid_rows=total_rows - len(items),
        total_units=sum(i.quantity for i in items),
        total_weight=round(sum(i.quantity * i.weight for i in items), 2),
        unique_codes=len(seen_codes),
    )

    return ImportPreview(profile=profile, is_valid=len(errors) == 0, items=items, errors=errors, warnings=warnings, summary=summary)
