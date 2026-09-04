"""Generador de templates de importacion CUBOX 2.0 (.xlsx) por perfil
(BOX/PALLET/PANEL/CUSTOM), Fase 3B.

Las columnas generadas aca deben coincidir EXACTAMENTE (mismos nombres, sin
importar mayusculas/espacios) con las que reconoce
core/import_items.py:_COLUMN_ALIASES -este modulo no reimplementa ningun
mapeo de dimensiones ni de orientacion, solo arma el workbook."""

import io

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from app.models.schemas import ItemType

TEMPLATE_VERSION = 1

_ORIENTATION_OPTIONS_BY_PROFILE: dict[ItemType, list[str]] = {
    ItemType.BOX: ["FREE", "UPRIGHT", "FIXED"],
    ItemType.PALLET: ["UPRIGHT", "FIXED"],
    ItemType.CUSTOM: ["FREE", "UPRIGHT", "FIXED"],
}

_HEADERS_BY_PROFILE: dict[ItemType, list[str]] = {
    ItemType.BOX: [
        "Code", "Quantity", "Length", "Width", "Height", "Weight",
        "Description", "Orientation", "Stackable", "Max Stack Weight", "Group", "Priority", "Delivery Sequence",
    ],
    ItemType.PALLET: [
        # Orientation es OPCIONAL (Fase 6): si se omite, se usa el Floor
        # Rotation del plan (wizard) o UPRIGHT de sistema -ver import_items.py.
        "Code", "Quantity", "Length", "Width", "Height", "Weight",
        "Description", "Orientation", "Stackable", "Max Stack Weight", "Group", "Priority", "Delivery Sequence",
    ],
    ItemType.PANEL: [
        "Code", "Quantity", "Width", "Height", "Thickness", "Weight",
        "Description", "System", "Group", "Stackable", "Max Stack Weight", "Priority", "Delivery Sequence",
    ],
    ItemType.CUSTOM: [
        "Code", "Quantity", "Length", "Width", "Height", "Weight", "Orientation",
        "Description", "Stackable", "Max Stack Weight", "Group", "Priority", "Delivery Sequence",
    ],
}

_COLUMN_NOTES: dict[str, str] = {
    "Code": "Identificador unico del item. Obligatorio.",
    "Quantity": "Cantidad de unidades identicas. Obligatorio, entero mayor a 0.",
    "Length": "Dimension generica 'largo', en mm. Obligatorio, mayor a 0.",
    "Width": "Dimension generica 'ancho', en mm. Obligatorio, mayor a 0.",
    "Height": "Dimension generica 'alto', en mm. Obligatorio, mayor a 0.",
    "Thickness": "Espesor del panel/vidrio, en mm. Obligatorio, mayor a 0. "
    "La cara Width x Height jamas puede quedar como base.",
    "Weight": "Peso unitario en kg. Obligatorio, mayor a 0.",
    "Description": "Texto libre. Opcional.",
    "Orientation": "FREE, UPRIGHT o FIXED. FREE = sin restriccion (6 orientaciones posibles). "
    "UPRIGHT = Height siempre permanece vertical. FIXED = sin cambios de orientacion.",
    "Stackable": "Yes/No. Si se omite, se asume No para este perfil.",
    "Max Stack Weight": "Peso maximo (kg) que puede soportar encima de este item. Vacio = sin limite.",
    "Group": "Etiqueta de agrupamiento. Opcional.",
    "System": "Sistema/linea de producto (uso tipico en ventanas). Opcional.",
    "Priority": "Prioridad de carga (numero entero). Vacio = 0 (prioridad media).",
    "Delivery Sequence": "Orden de entrega/parada. Opcional, numerico.",
}

_EXAMPLE_VALUES: dict[str, object] = {
    "Code": "EXAMPLE-001",
    "Quantity": 1,
    "Length": 600,
    "Width": 400,
    "Height": 300,
    "Thickness": 100,
    "Weight": 25,
    "Description": "Example row - delete before importing",
    "Orientation": "FREE",
    "Stackable": "No",
}

# Fase 6: PALLET solo admite UPRIGHT/FIXED (nunca FREE) -la nota generica y
# el valor de ejemplo de "Orientation" de arriba no aplican para ese perfil,
# asi que se sobreescriben puntualmente aca en vez de bifurcar todo el dict.
_ORIENTATION_NOTE_OVERRIDE_BY_PROFILE: dict[ItemType, str] = {
    ItemType.PALLET: "Opcional. UPRIGHT o FIXED (nunca FREE: un pallet jamas se para de canto). Si se omite, "
    "se usa el Floor Rotation configurado en el plan (Handling Rules); si el plan tampoco lo define, UPRIGHT. "
    "UPRIGHT = puede rotar 90 grados en el piso (Length x Width), Height siempre vertical. "
    "FIXED = sin cambios de orientacion.",
}
_ORIENTATION_EXAMPLE_OVERRIDE_BY_PROFILE: dict[ItemType, str] = {
    ItemType.PALLET: "UPRIGHT",
}


def build_import_template(profile: ItemType) -> bytes:
    headers = _HEADERS_BY_PROFILE[profile]

    workbook = Workbook()
    items_sheet = workbook.active
    items_sheet.title = "Items"
    items_sheet.append(headers)

    _add_dropdown(items_sheet, headers, "Orientation", _ORIENTATION_OPTIONS_BY_PROFILE.get(profile))
    _add_dropdown(items_sheet, headers, "Stackable", ["Yes", "No"])

    _build_instructions_sheet(workbook, profile, headers)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _add_dropdown(sheet, headers: list[str], column_name: str, options: list[str] | None) -> None:
    if not options or column_name not in headers:
        return
    col_index = headers.index(column_name) + 1
    validation = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
    sheet.add_data_validation(validation)
    column_letter = sheet.cell(row=1, column=col_index).column_letter
    validation.add(f"{column_letter}2:{column_letter}1000")


def _build_instructions_sheet(workbook: Workbook, profile: ItemType, headers: list[str]) -> None:
    sheet = workbook.create_sheet("Instructions")
    sheet.append(["CUBOX Import Template"])
    sheet.append([f"Profile: {profile.value.upper()}"])
    sheet.append([f"Template Version: {TEMPLATE_VERSION}"])
    sheet.append([])
    sheet.append(["Column", "Notes"])
    orientation_note = _ORIENTATION_NOTE_OVERRIDE_BY_PROFILE.get(profile, _COLUMN_NOTES.get("Orientation", ""))
    for header in headers:
        note = orientation_note if header == "Orientation" else _COLUMN_NOTES.get(header, "")
        sheet.append([header, note])

    sheet.append([])
    sheet.append(["Example row (for reference only - do not import as real data):"])
    sheet.append(headers)
    orientation_example = _ORIENTATION_EXAMPLE_OVERRIDE_BY_PROFILE.get(profile, _EXAMPLE_VALUES.get("Orientation", ""))
    sheet.append([orientation_example if h == "Orientation" else _EXAMPLE_VALUES.get(h, "") for h in headers])
