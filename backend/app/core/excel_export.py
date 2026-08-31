"""Exportacion a Excel del resultado de cubicaje actual (seccion 34 de V2)."""

import io

from openpyxl import Workbook

from app.models.schemas import PackingResult


def build_export_workbook(state: PackingResult) -> bytes:
    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    summary.append(["Container", state.container.name])
    summary.append(["Loaded", state.metrics.loaded_pieces])
    summary.append(["Unloaded", state.metrics.unloaded_pieces])
    summary.append(["Volume Utilization %", state.metrics.used_volume_pct])
    summary.append(["Floor Utilization %", state.metrics.floor_utilization_pct])
    summary.append(["Total Weight (kg)", state.metrics.total_weight])
    summary.append(["Max Payload (kg)", state.metrics.max_payload])
    summary.append(["Weight Utilization %", state.metrics.weight_utilization_pct])
    summary.append(["Number of Groups", state.metrics.number_of_groups])
    summary.append(["Number of Systems", state.metrics.number_of_systems])

    packing_list = wb.create_sheet("Packing List")
    packing_list.append(
        [
            "Load Order",
            "Unload Order",
            "Code",
            "Description",
            "Width",
            "Height",
            "Thickness",
            "Weight",
            "Group",
            "System",
            "Priority",
        ]
    )
    load_order = {piece_id: i + 1 for i, piece_id in enumerate(state.load_sequence)}
    unload_order = {piece_id: i + 1 for i, piece_id in enumerate(state.unload_sequence)}
    for p in state.placed:
        packing_list.append(
            [
                load_order.get(p.id, ""),
                unload_order.get(p.id, ""),
                p.code,
                p.description,
                p.source_width,
                p.source_height,
                p.source_thickness,
                p.weight,
                p.group,
                p.system,
                p.priority,
            ]
        )

    unloaded_sheet = wb.create_sheet("Unloaded Items")
    unloaded_sheet.append(["Code", "Description", "Width", "Height", "Thickness", "Weight", "Reason"])
    for u in state.unloaded:
        unloaded_sheet.append([u.code, u.description, u.width, u.height, u.thickness, u.weight, u.reason])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
