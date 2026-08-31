"""Importacion de la lista de ventanas desde un archivo Excel (.xlsx).

Columnas esperadas (en cualquier orden, encabezados case-insensitive):
Code, Description, Width, Height, Thickness, Weight, Quantity, System,
Group, Stackable, Priority. MaxStackWeight y DeliverySequence son opcionales
(vacio = sin limite / sin definir) - un Excel de V1/V2 sin esas columnas
sigue importando igual.
"""

import io

from openpyxl import load_workbook

from app.models.schemas import WindowItem

REQUIRED_COLUMNS = ["code", "width", "height", "thickness", "weight", "quantity"]

COLUMN_ALIASES = {
    "code": "code",
    "description": "description",
    "width": "width",
    "height": "height",
    "thickness": "thickness",
    "weight": "weight",
    "quantity": "quantity",
    "system": "system",
    "group": "group",
    "stackable": "stackable",
    "priority": "priority",
    "maxstackweight": "max_stack_weight",
    "max_stack_weight": "max_stack_weight",
    "deliverysequence": "delivery_sequence",
    "delivery_sequence": "delivery_sequence",
    "stop": "delivery_sequence",
}

TRUE_VALUES = {"yes", "y", "true", "1", "si", "sí", "x"}


def _parse_stackable(value) -> bool:
    if value is None:
        return True
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in TRUE_VALUES


def _parse_row(headers: list[str], row: tuple, row_number: int) -> WindowItem:
    values = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}

    missing = [c for c in REQUIRED_COLUMNS if values.get(c) in (None, "")]
    if missing:
        raise ValueError(f"Fila {row_number}: faltan columnas obligatorias {missing}")

    return WindowItem(
        code=str(values["code"]).strip(),
        description=str(values.get("description") or "").strip(),
        width=float(values["width"]),
        height=float(values["height"]),
        thickness=float(values["thickness"]),
        weight=float(values["weight"]),
        quantity=int(values["quantity"]),
        system=str(values.get("system") or "").strip(),
        group=str(values.get("group") or "").strip(),
        stackable=_parse_stackable(values.get("stackable")),
        priority=int(values.get("priority") or 0),
        max_stack_weight=float(values["max_stack_weight"]) if values.get("max_stack_weight") not in (None, "") else None,
        delivery_sequence=int(values["delivery_sequence"]) if values.get("delivery_sequence") not in (None, "") else None,
    )


def parse_excel(file_bytes: bytes) -> list[WindowItem]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError("El archivo Excel esta vacio")

    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]
    headers = [COLUMN_ALIASES.get(h, h) for h in headers]

    missing_cols = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing_cols:
        raise ValueError(f"Columnas obligatorias faltantes en el Excel: {missing_cols}")

    items: list[WindowItem] = []
    for row_number, row in enumerate(rows_iter, start=2):
        if row is None or all(v is None for v in row):
            continue
        items.append(_parse_row(headers, row, row_number))

    if not items:
        raise ValueError("El Excel no contiene filas de datos")

    return items
